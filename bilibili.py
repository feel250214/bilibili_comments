import jieba
import pandas as pd
import requests, re, json, random, base64, hashlib, string, urllib, time
import re
import urllib.parse
from bs4 import BeautifulSoup
import requests.cookies
import xlwt
import openpyxl

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def word_frequency_xlsx(xlsx_name='bilibili_comment.xlsx'):
    df = pd.read_excel(xlsx_name)  # 读取 Excel 文件
    txt_name = "xlsx_to_txt.txt"
    m = 1
    # with open(txt_name, 'w', encoding="utf-8") as fp:  # 清空文件
    #     pass

    while m < len(df):
        data = df.iloc[m, 0]  # 直接获取数据
        with open(txt_name, 'a', encoding="utf-8") as fp:  # 追加模式
            fp.writelines(str(data) + '\n')  # 每行数据换行
        m += 1  # 递增索引
    word_frequency_txt(txt_name)


def word_frequency_txt(txt_name='bilibili_comments.txt'):
    """
    统计词频
    """
    with open('baidu_stopwords.txt', 'r', encoding='utf-8') as f:  # 读入停用词文件
        stopwords = set(line.strip() for line in f)

    # 读入文件
    with open(txt_name, encoding="utf-8") as f:
        text = f.read()

    ls = jieba.lcut(text, cut_all=True)  # 分词
    # 统计词频
    counts = {}
    for i in ls:
        if len(i) > 1:
            counts[i] = counts.get(i, 0) + 1
    for word in stopwords:  # 去掉停用词
        counts.pop(word, 0)
    ls1 = sorted(counts.items(), key=lambda x: x[1], reverse=True)  # 词频排序
    # 输出
    print(ls1)
    with open('bili_word_frequency.txt', 'w', encoding="utf-8") as fp:
        fp.writelines(str(ls1) + '\n')

def save_to_txt(comment_list, txt_name='bilibili_comments.txt'):
    """
    保存评论到txt，不保存性别
    """
    with open(txt_name, 'w', encoding="utf-8") as f:  # 清空文件
        pass
    with open(txt_name, 'a', encoding='utf-8') as f:  # 追加模式
        for comment in comment_list:
            f.write(comment + '\n')  # 每行数据换行


def _get_oid(url):
    """
    从视频链接中获取 oid（视频 ID），用于获取评论
    """
    oid_pattern = r'"aid":(\d+)'
    response = request_bili(url)
    response.raise_for_status()
    response.encoding = response.apparent_encoding

    html_content = response.text
    match = re.search(oid_pattern, html_content)
    if match:
        oid = match.group(1)
        return oid
    else:
        print("fail matching oid")
        return None


def get_random_user_agent():
    return random.choice(user_agents)


def get_timestamp():
    """
    获取时间码
    """
    return time.time()


def names_save_to_excel(video_urls_list, video_names_list):
    """
    保存视频信息（url 和 name）到 excel
    """
    book = openpyxl.Workbook()

    sheet = book.create_sheet('各视频名称和对应url', index=0)

    sheet.cell(1, 1).value = '名字'
    sheet.cell(1, 2).value = 'url'
    m = 2
    for name, url in zip(video_names_list, video_urls_list):
        name = ILLEGAL_CHARACTERS_RE.sub(r'', name)
        sheet.cell(m, 1).value = name
        url = ILLEGAL_CHARACTERS_RE.sub(r'', url)
        sheet.cell(m, 2).value = url
        m = m + 1
    book.save('bilibili_videos_names.xlsx')


def comments_save_to_excel(comments_list, sexs_list, excel_name='bilibili_comments.xlsx'):
    """
    将视频评论保存到 Excel
    """
    book = openpyxl.Workbook()

    sheet = book.create_sheet('各视频评论', index=0)

    sheet.cell(1, 1).value = '评论'
    sheet.cell(1, 2).value = '性别'

    m = 2
    for comment, sex in zip(comments_list, sexs_list):
        comment = ILLEGAL_CHARACTERS_RE.sub(r'', comment)
        sheet.cell(m, 1).value = comment
        sex = ILLEGAL_CHARACTERS_RE.sub(r'', sex)
        sheet.cell(m, 2).value = sex
        m = m + 1
    book.save(excel_name)


def text_to_encoded(keyword_list):
    """
    将关键词转换为URL编码（好像不用也行）
    """
    encoded_list = []
    for keyword in keyword_list:
        encoded_list.append(urllib.parse.quote(keyword))
    return encoded_list


def request_bili(url, params=None, print_url=True):
    """
    嘗試訪問網站
    """
    time.sleep(random.random())
    response = None
    n = 0

    while True:
        try:
            if params is not None:
                response = requests.get(url, params, headers=headers)
            else:
                response = requests.get(url, headers=headers)
            n += 1
            if response.status_code == 200:
                if print_url:
                    print(url + "访问成功")
                return response
            else:
                n += 1
                continue
        except requests.exceptions.RequestException as e:
            if n > 20:
                print(f"status_code={response.status_code} when req '{url}'")
                print(e)
                break
            continue


def get_videos_url(encodes, is_save_to_excel=True):
    """
    获得各个视频的url
    """
    filtered_urls = []
    filtered_names = []
    video_names_list = []
    video_urls_list = []
    for encoded in encodes:
        o = 0
        for page in range(1, 25):   # page页数最多25
            if o == 0:
                url = 'https://search.bilibili.com/all?keyword=' + str(encoded) + '&search_source=1'
            else:
                url = 'https://search.bilibili.com/all?keyword=' + str(encoded) + '&search_source=1&page=' + str(
                    page) + '&o=' + str(o)
            html = request_bili(url)
            soup = BeautifulSoup(html.text, 'lxml')
            a = soup.find_all('div', class_='bili-video-card__info--right')
            for i in a:
                href = i.find('a')['href']
                name = i.find('h3', class_='bili-video-card__info--tit').get_text()
                video_names_list.append(name)
                video_urls_list.append(href)
            o += 42

        # 筛选以 //www.bilibili.com/ 开头的 URL 并同步删除对应名字，去除直播和课程视频
    for url, name in zip(video_urls_list, video_names_list):
        if url.startswith('//www.bilibili.com/'):
            filtered_urls.append("https:" + url)
            filtered_names.append(name)
    print("去除直播和课程视频后共有" + str(len(filtered_urls)) + "条视频")
    if is_save_to_excel:
        names_save_to_excel(video_urls_list, video_names_list)
    return filtered_urls, filtered_names


def _get_pagiantion_str(next_offset):
    """
    生成分页参数的字符串形式
    """
    pagination = {"offset": ''}
    pagination['offset'] = next_offset
    pagination_str = json.dumps(pagination).replace(' ', '')
    return pagination_str


def _bili_w_rid(req: dict):
    """
    用于生成w_rid（请求参数）
    """
    offset = "ea1db124af3c7062474693fa704f4ff8"
    page_str = req['pagination_str']
    encoded_str = urllib.parse.quote(page_str).replace('%25', '%')
    # print(f"str = {page_str}\nencoded = {encoded_str}")
    copy = req.copy()
    copy['pagination_str'] = encoded_str
    string_list = [f"{k}={v}" for k, v in copy.items()]
    L = "&".join(sorted(string_list))
    string_joint = L + offset
    # print(f"L+offset = {string_joint}")
    MD5 = hashlib.md5()
    MD5.update(string_joint.encode('utf-8'))
    w_rid = MD5.hexdigest()
    return w_rid


def get_comments(url, mode=3, pages=10):
    """
    获取视频评论及评论者性别
    """
    global response
    comment_url = 'https://api.bilibili.com/x/v2/reply/wbi/main'
    params_template = {
        "oid": "",
        "mode": 3,
        "pagination_str": '{"offset":""}',
        "plat": 1,
        "type": 1,
        "web_location": 1315875,
        "wts": 0,
    }
    first_params_addition = {
        "seek_rpid": "",
    }
    if pages < 1:
        return []
    comments_list = []
    sexs_list = []
    params_template = params_template.copy()
    oid = _get_oid(url)
    if oid is None:
        print('oid failed')
        return comments_list
    params_template['mode'] = mode
    params_template['oid'] = oid
    params = params_template.copy()
    # first
    params.update(first_params_addition)

    page = 1
    last = []
    while True:
        params['wts'] = get_timestamp()
        params['w_rid'] = _bili_w_rid(params)

        try:
            response = request_bili(comment_url, params=params, print_url=False)
        except:
            pass
        comments = []
        sexs = []
        try:
            replies = response.json()['data']['replies']
            for reply in replies:
                comments.append(reply['content']['message'])
                sexs.append(reply['member']['sex'])
        except:
            # print('comments reply failed')
            # print(response.text)
            # print(params)
            pass
        if comments == last:
            break
        last = comments
        comments_list += comments
        sexs_list += sexs

        try:
            next_offset = response.json()['data']['cursor']['pagination_reply']['next_offset']
        except:
            # print('pagination reply failed')
            break
        page += 1
        if page > pages:
            break

        # prepare next page_str
        params = params_template.copy()
        params['pagination_str'] = _get_pagiantion_str(next_offset)
        # print(params)
        # print(params['pagination_str'])

    return comments_list, sexs_list


if __name__ == '__main__':
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15",
        # "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QP1A.190711.020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Mobile Safari/537.36",
        # "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Gecko/20100101 Firefox/89.0",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
        # "Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Linux; Android 10; SM-G975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
        # "Mozilla/5.0 (iPad; CPU OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/77.0.4054.203",
        # "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 10; SM-G980F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Vivaldi/4.0",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 9; SM-G960F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Brave/1.26.77",
        # "Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/91.0.4472.124 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Linux; Android 10; SM-N975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 11; Pixel 4 XL) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/77.0.4054.203",
        # "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        # "Mozilla/5.0 (Linux; Android 10; SM-G970F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Vivaldi/4.0",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 9; SM-G965F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Brave/1.26.77",
        # "Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/91.0.4472.124 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Linux; Android 10; SM-N960F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 11; Pixel 4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/77.0.4054.203",
        # "Mozilla/5.0 (X11; Fedora; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        # "Mozilla/5.0 (Linux; Android 10; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Vivaldi/4.0",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 9; SM-G960U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Brave/1.26.77",
        # "Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/91.0.4472.124 Mobile/15E148 Safari/604.1",
        # "Mozilla/5.0 (Linux; Android 10; SM-N970F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 11; Pixel 3 XL) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/77.0.4054.203",
        # "Mozilla/5.0 (X11; Debian; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        # "Mozilla/5.0 (Linux; Android 10; SM-G975U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
        # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Vivaldi/4.0",
        # "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        # "Mozilla/5.0 (Linux; Android 9; SM-G965U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
    ]  # User-agent pool
    headers = {
        "User-Agent": get_random_user_agent(),
        # "Cookie": "更换新Cookie"
    }
    comments_list = []
    sexs_list = []
    comments = []
    sexs = []

    keyword = ["新能源"]
    encodes = text_to_encoded(keyword)
    [video_urls_list, video_names_list] = get_videos_url(encodes, False)
    for video_url in video_urls_list:
        [comments, sexs] = get_comments(video_url, 3, 2000)
        comments_list += comments
        sexs_list += sexs
        print(video_url + "共抓取" + str(len(comments)) + "条评论")
        print("现在一共抓取" + str(len(comments_list)) + "条评论")
    comments_save_to_excel(comments_list, sexs_list)
    # word_frequency_xlsx()
